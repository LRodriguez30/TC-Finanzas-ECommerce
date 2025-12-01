import customtkinter as ctk
from ..theme_manager import get_color
import pandas as pd
import numbers
from backend.DAOs.CuentasDAO import CuentaDAO
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

class FinancialEcommercePage(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color="transparent")
        
        # Header
        self.header_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.header_frame.pack(fill="x", pady=(0, 20))
        
        title = ctk.CTkLabel(self.header_frame, text="Análisis Financiero de Ecommerce", font=ctk.CTkFont(size=22, weight="bold"), text_color="#111827")
        title.pack(side="left")
        
        # Header Buttons
        btn_frame = ctk.CTkFrame(self.header_frame, fg_color="transparent")
        btn_frame.pack(side="right")
        
        ctk.CTkButton(btn_frame, text="Exportar Reporte", command=self.export_report, fg_color=get_color('success')).pack(side="left", padx=6)
        ctk.CTkButton(btn_frame, text="Interpretar", command=self.show_interpretations, fg_color=get_color('accent')).pack(side="left", padx=6)
        
        # --- Pestañas ---
        self.tabview = ctk.CTkTabview(self, segmented_button_selected_color="#F97316", segmented_button_selected_hover_color="#D97706")
        self.tabview.pack(fill="both", expand=True)
        
        self.tabs = {
            "Cuentas": self.tabview.add("Cuentas"),
            "Balance": self.tabview.add("Balance General"),
            "Resultados": self.tabview.add("Estado de Resultados"),
            "Origen": self.tabview.add("Origen y Aplicación"),
            "Razones": self.tabview.add("Razones Financieras"),
            "Flujos": self.tabview.add("Flujos de Efectivo"),
            "Graficos": self.tabview.add("Gráficos"),
            "Proforma": self.tabview.add("Proforma")
        }
        
        
        # Estado inicial vacío para pestañas no implementadas
        for tab_name, tab in self.tabs.items():
            if tab_name in ["Resultados", "Balance", "Origen", "Razones", "Flujos", "Graficos", "Proforma"]:
                continue

        # Cargar datos
        self.load_platform_accounts()
        self.load_income_statement()
        self.load_balance_sheet()
        self.load_cash_flow()
        self.load_ratios()
        self.load_cash_flow_statements()
        self.load_charts()
        self.render_proforma(self.tabs["Proforma"])
    
    def load_platform_accounts(self):
        """Carga las cuentas de la plataforma en la pestaña de Cuentas"""
        cuentas_tab = self.tabs["Cuentas"]
        
        # Limpiar la pestaña
        for widget in cuentas_tab.winfo_children():
            widget.destroy()
        
        # Header con título
        header_frame = ctk.CTkFrame(cuentas_tab, fg_color="transparent")
        header_frame.pack(fill="x", padx=20, pady=(20, 10))
        
        ctk.CTkLabel(
            header_frame,
            text="Cuentas de la Plataforma",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color="#111827"
        ).pack(side="left")
        
        # Obtener todas las cuentas de la plataforma
        all_balance = CuentaDAO.obtener_cuentas_balance_plataforma()
        all_results = CuentaDAO.obtener_cuentas_resultados_plataforma()
        all_platform_accounts = all_balance + all_results
        
        if not all_platform_accounts:
            ctk.CTkLabel(
                cuentas_tab,
                text="No se encontraron cuentas de la plataforma",
                text_color="#6B7280"
            ).pack(pady=20)
            return
        
        # Crear tabla de cuentas
        table_frame = ctk.CTkScrollableFrame(cuentas_tab, fg_color="transparent")
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Headers
        headers = ["Código", "Nombre", "Tipo", "Saldo", "Sistema"]
        header_frame = ctk.CTkFrame(table_frame, fg_color="#F3F4F6")
        header_frame.pack(fill="x", pady=(0, 5))
        
        for i, header in enumerate(headers):
            ctk.CTkLabel(
                header_frame,
                text=header,
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color="#374151",
                width=150 if i < 4 else 80
            ).grid(row=0, column=i, padx=10, pady=8, sticky="w")
        
        # Datos de cuentas
        tipos_cuenta = {
            1: "Activo Corriente",
            2: "Activo No Corriente",
            3: "Pasivo Corriente",
            4: "Pasivo No Corriente",
            5: "Capital",
            6: "Ingreso",
            7: "Gasto Operativo",
            8: "Gasto de Interés"
        }
        
        for cuenta in all_platform_accounts:
            row_frame = ctk.CTkFrame(table_frame, fg_color="transparent")
            row_frame.pack(fill="x", pady=2)
            
            # Código
            ctk.CTkLabel(
                row_frame,
                text=cuenta.codigo_cuenta,
                text_color="#111827",
                width=150
            ).grid(row=0, column=0, padx=10, pady=5, sticky="w")
            
            # Nombre
            ctk.CTkLabel(
                row_frame,
                text=cuenta.nombre_cuenta,
                text_color="#111827",
                width=150
            ).grid(row=0, column=1, padx=10, pady=5, sticky="w")
            
            # Tipo
            ctk.CTkLabel(
                row_frame,
                text=tipos_cuenta.get(cuenta.id_tipo_cuenta, "Desconocido"),
                text_color="#6B7280",
                width=150
            ).grid(row=0, column=2, padx=10, pady=5, sticky="w")
            
            # Saldo
            ctk.CTkLabel(
                row_frame,
                text=f"${cuenta.saldo_actual:,.2f}",
                text_color="#059669" if cuenta.saldo_actual >= 0 else "#EF4444",
                width=150
            ).grid(row=0, column=3, padx=10, pady=5, sticky="w")
            
            # Sistema
            sistema_text = "✓" if cuenta.es_cuenta_de_sistema else ""
            ctk.CTkLabel(
                row_frame,
                text=sistema_text,
                text_color="#F97316",
                width=80
            ).grid(row=0, column=4, padx=10, pady=5, sticky="w")

    def load_balance_sheet(self):
        # 1. Datos del Período Anterior (Hardcoded)
        prev_data_raw = [
            {"Cuenta": "ACTIVOS", "Monto": None}, # Header
            {"Cuenta": "Efectivo", "Monto": 0},
            {"Cuenta": "Mobiliario y accesorios", "Monto": 36000},
            {"Cuenta": "Vehículos", "Monto": 50000},
            {"Cuenta": "Depreciación de mobiliario y accesorios", "Monto": 0},
            {"Cuenta": "Depreciación de vehículos", "Monto": 0},
            {"Cuenta": "Total Activos", "Monto": 86000},
            {"Cuenta": "PASIVOS", "Monto": None}, # Header
            {"Cuenta": "Total Pasivos", "Monto": 0},
            {"Cuenta": "CAPITAL", "Monto": None}, # Header
            {"Cuenta": "Capital Social", "Monto": 86000},
            {"Cuenta": "Utilidades Retenidas", "Monto": 0},
            {"Cuenta": "Total Capital", "Monto": 86000},
            {"Cuenta": "Total Pasivo + Capital", "Monto": 86000},
            {"Cuenta": "Diferencia", "Monto": 0}
        ]

        # Calcular Vertical % (Base = Total Activos = 86000)
        total_activos_prev = 86000
        for row in prev_data_raw:
            if row["Monto"] is not None:
                row["Vertical %"] = (row["Monto"] / total_activos_prev * 100) if total_activos_prev else 0.0
            else:
                row["Vertical %"] = None

        df_prev = pd.DataFrame(prev_data_raw)

        # 2. Datos del Período Actual (Desde BD)
        current_accounts = CuentaDAO.obtener_cuentas_balance_plataforma()
        
        curr_data_map = {}
        curr_activos = 0
        curr_pasivos = 0
        curr_capital = 0
        
        # Clasificación simple basada en IDs o nombres (Heurística)
        # Tipos: 1,2 = Activo? 5 = Capital? 
        # Nombres: 'Efectivo', 'Mobiliario', 'Vehiculos' -> Activos
        # 'Capital social', 'Utilidades retenidas' -> Capital
        # 'Renta', 'Gasto' -> No deberían estar aquí (son Resultados)
        
        for acc in current_accounts:
            name = acc.nombre_cuenta
            val = float(acc.saldo_actual)
            curr_data_map[name] = val
            
            # Sumar a totales (simplificado por nombre/tipo)
            # Asumimos que la BD tiene los tipos correctos.
            # Ajuste manual para totales basado en nombres conocidos
            if name in ["Efectivo", "Mobiliario y accesorios", "Vehiculos", "Vehículos"]:
                curr_activos += val
            elif "Depreciacion" in name or "Depreciación" in name:
                 # Depreciaciones suelen restar, pero si vienen positivas en BD, hay que ver.
                 # En el ejemplo del usuario, se restan: 0 + 36000 + 50000 - 0 - 0
                 # Asumiremos que se restan si son depreciaciones
                 curr_activos -= val 
            elif name in ["Capital social", "Utilidades retenidas"]:
                curr_capital += val
            # Pasivos?
        
        # Totales calculados
        curr_pasivo_mas_capital = curr_pasivos + curr_capital
        curr_diferencia = curr_activos - curr_pasivo_mas_capital

        # Mapear a estructura
        curr_data_list = []
        
        # Usar estructura de prev para orden
        for row in prev_data_raw:
            key = row["Cuenta"]
            
            if key in ["ACTIVOS", "PASIVOS", "CAPITAL"]: # Headers
                curr_data_list.append({"Cuenta": key, "Monto": None, "Vertical %": None})
                continue
                
            # Totales calculados
            if key == "Total Activos":
                val = curr_activos
            elif key == "Total Pasivos":
                val = curr_pasivos
            elif key == "Total Capital":
                val = curr_capital
            elif key == "Total Pasivo + Capital":
                val = curr_pasivo_mas_capital
            elif key == "Diferencia":
                val = curr_diferencia
            else:
                # Cuentas individuales
                # Normalizar keys - buscar con case-insensitive
                lookup_key = key
                val = 0.0
                
                # Intentar búsqueda exacta primero
                if key in curr_data_map:
                    val = curr_data_map[key]
                else:
                    # Intentar variaciones comunes
                    if key == "Vehículos" and "Vehiculos" in curr_data_map:
                        val = curr_data_map["Vehiculos"]
                    elif "Depreciación" in key:
                        simple = key.replace("Depreciación", "Depreciacion")
                        val = curr_data_map.get(simple, 0.0)
                    else:
                        # Búsqueda case-insensitive
                        key_lower = key.lower()
                        for db_key, db_val in curr_data_map.items():
                            if db_key.lower() == key_lower:
                                val = db_val
                                break

            vert = (val / curr_activos * 100) if (curr_activos and val is not None) else 0.0
            if val is None: vert = None
            
            curr_data_list.append({"Cuenta": key, "Monto": val, "Vertical %": vert})

        df_curr = pd.DataFrame(curr_data_list)

        # 3. Renderizar UI
        tab = self.tabs["Balance"]
        for widget in tab.winfo_children():
            widget.destroy()

        # Contenedor principal
        container = ctk.CTkFrame(tab, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=10, pady=10)

        # Frame Izquierdo (Anterior)
        left_frame = ctk.CTkFrame(container, fg_color="transparent")
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))

        ctk.CTkLabel(left_frame, text="Período Anterior", font=ctk.CTkFont(size=16, weight='bold'), text_color="#111827").pack(anchor='w', pady=(0, 5))
        self.render_table(left_frame, df_prev)

        # Frame Derecho (Actual)
        right_frame = ctk.CTkFrame(container, fg_color="transparent")
        right_frame.pack(side="left", fill="both", expand=True, padx=(10, 0))

        ctk.CTkLabel(right_frame, text="Período Actual", font=ctk.CTkFont(size=16, weight='bold'), text_color="#111827").pack(anchor='w', pady=(0, 5))
        self.render_table(right_frame, df_curr)
        
        # 4. Análisis Horizontal (Abajo)
        horizontal_data = []
        for i in range(len(prev_data_raw)):
            row_prev = prev_data_raw[i]
            row_curr = curr_data_list[i]
            
            key = row_prev["Cuenta"]
            val_prev = row_prev["Monto"]
            val_curr = row_curr["Monto"]
            
            if val_prev is None or val_curr is None:
                continue # Skip headers
                
            var_abs = val_curr - val_prev
            var_pct = (var_abs / val_prev * 100) if val_prev != 0 else 0.0
            
            horizontal_data.append({
                "Cuenta": key,
                "Año Base": val_prev,
                "Año Actual": val_curr,
                "Variación $": var_abs,
                "Variación %": var_pct
            })
            
        df_horizontal = pd.DataFrame(horizontal_data)
        ctk.CTkLabel(tab, text="Análisis Horizontal (Variación)", font=ctk.CTkFont(size=16, weight='bold'), text_color="#111827").pack(anchor='w', padx=10, pady=(20, 5))
        self.render_table(tab, df_horizontal)

    def load_cash_flow(self):
        """Calculate and display Sources and Uses (Cash Flow Statement)"""
        # We need the balance sheet data from both periods
        # Reuse the data structures from load_balance_sheet
        
        # Previous period data
        prev_data_raw = [
            {"Cuenta": "ACTIVOS", "Monto": None, "Tipo": "Header"},
            {"Cuenta": "Efectivo", "Monto": 0, "Tipo": "Activo"},
            {"Cuenta": "Mobiliario y accesorios", "Monto": 36000, "Tipo": "Activo"},
            {"Cuenta": "Vehículos", "Monto": 50000, "Tipo": "Activo"},
            {"Cuenta": "Depreciación de mobiliario y accesorios", "Monto": 0, "Tipo": "Activo"},
            {"Cuenta": "Depreciación de vehículos", "Monto": 0, "Tipo": "Activo"},
            {"Cuenta": "PASIVOS", "Monto": None, "Tipo": "Header"},
            {"Cuenta": "CAPITAL", "Monto": None, "Tipo": "Header"},
            {"Cuenta": "Capital Social", "Monto": 86000, "Tipo": "Capital"},
            {"Cuenta": "Utilidades Retenidas", "Monto": 0, "Tipo": "Capital"}
        ]
        
        # Current period data (fetch from DB)
        current_accounts = CuentaDAO.obtener_cuentas_balance_plataforma()
        curr_data_map = {}
        
        for acc in current_accounts:
            name = acc.nombre_cuenta
            val = float(acc.saldo_actual)
            tipo = None
            
            # Classify by type
            if name in ["Efectivo", "Mobiliario y accesorios", "Vehiculos", "Vehículos"] or "Depreciacion" in name or "Depreciación" in name:
                tipo = "Activo"
            elif name in ["Capital social", "Utilidades retenidas"]:
                tipo = "Capital"
            else:
                tipo = "Pasivo"  # Default for unknown
                
            curr_data_map[name] = {"Monto": val, "Tipo": tipo}
        
        # Calculate Sources and Uses
        sources = []
        uses = []
        
        for row in prev_data_raw:
            cuenta = row["Cuenta"]
            monto_prev = row["Monto"]
            tipo = row["Tipo"]
            
            if monto_prev is None or tipo == "Header":
                continue
                
            # Find current value (with improved matching)
            monto_curr = 0.0
            tipo_curr = tipo
            
            # Helper function to normalize account names
            def normalize_name(name):
                """Remove accents and normalize for comparison"""
                import unicodedata
                # Remove accents
                normalized = unicodedata.normalize('NFD', name)
                normalized = ''.join(c for c in normalized if unicodedata.category(c) != 'Mn')
                return normalized.lower().strip()
            
            # Try exact match first
            if cuenta in curr_data_map:
                monto_curr = curr_data_map[cuenta]["Monto"]
                tipo_curr = curr_data_map[cuenta]["Tipo"]
            else:
                # Try normalized matching (handles accents, case, etc.)
                cuenta_normalized = normalize_name(cuenta)
                for db_key, db_data in curr_data_map.items():
                    if normalize_name(db_key) == cuenta_normalized:
                        monto_curr = db_data["Monto"]
                        tipo_curr = db_data["Tipo"]
                        break
            
            diff = monto_curr - monto_prev
            
            if diff == 0 or abs(diff) < 0.01:  # Skip if no change (with small tolerance for floating point)
                continue
                
            # Sources and Uses logic:
            # - Asset decrease = Source
            # - Asset increase = Use
            # - Liability/Capital increase = Source
            # - Liability/Capital decrease = Use
            
            if tipo_curr == "Activo":
                if diff < 0:
                    sources.append({"Cuenta": cuenta, "Monto": abs(diff)})
                elif diff > 0:
                    uses.append({"Cuenta": cuenta, "Monto": abs(diff)})
            else:  # Pasivo or Capital
                if diff > 0:
                    sources.append({"Cuenta": cuenta, "Monto": abs(diff)})
                elif diff < 0:
                    uses.append({"Cuenta": cuenta, "Monto": abs(diff)})
        
        # Create DataFrames
        df_sources = pd.DataFrame(sources) if sources else pd.DataFrame(columns=["Cuenta", "Monto"])
        df_uses = pd.DataFrame(uses) if uses else pd.DataFrame(columns=["Cuenta", "Monto"])
        
        # Render UI (side-by-side like in financial.py)
        tab = self.tabs["Origen"]
        for widget in tab.winfo_children():
            widget.destroy()
        
        # Sources frame (left)
        sources_frame = ctk.CTkScrollableFrame(tab, label_text="Fuentes", fg_color=get_color('card_bg'))
        sources_frame.pack(side="left", fill="both", expand=True, padx=5, pady=10)
        
        if not df_sources.empty:
            for _, row in df_sources.iterrows():
                ctk.CTkLabel(sources_frame, text=f"{row['Cuenta']}: ${row['Monto']:,.2f}", text_color="#374151").pack(anchor="w", pady=2)
        else:
            ctk.CTkLabel(sources_frame, text="No hay fuentes de efectivo", text_color="#9CA3AF").pack(pady=10)
        
        # Uses frame (right)
        uses_frame = ctk.CTkScrollableFrame(tab, label_text="Usos", fg_color=get_color('card_bg'))
        uses_frame.pack(side="right", fill="both", expand=True, padx=5, pady=10)
        
        if not df_uses.empty:
            for _, row in df_uses.iterrows():
                ctk.CTkLabel(uses_frame, text=f"{row['Cuenta']}: ${row['Monto']:,.2f}", text_color="#374151").pack(anchor="w", pady=2)
        else:
            ctk.CTkLabel(uses_frame, text="No hay usos de efectivo", text_color="#9CA3AF").pack(pady=10)

    def load_ratios(self):
        """Calculate and display financial ratios for both periods"""
        # Fetch current period data
        balance_accounts = CuentaDAO.obtener_cuentas_balance_plataforma()
        income_accounts = CuentaDAO.obtener_cuentas_resultados_plataforma()
        
        # Helper to get account value by name
        def get_account_value(accounts_list, account_name):
            import unicodedata
            def normalize(name):
                normalized = unicodedata.normalize('NFD', name)
                normalized = ''.join(c for c in normalized if unicodedata.category(c) != 'Mn')
                return normalized.lower().strip()
            
            target_normalized = normalize(account_name)
            for acc in accounts_list:
                if normalize(acc.nombre_cuenta) == target_normalized:
                    return float(acc.saldo_actual)
            return 0.0
        
        # Calculate totals for CURRENT period
        curr_activos = 0
        curr_pasivos = 0
        curr_capital = 0
        
        for acc in balance_accounts:
            val = float(acc.saldo_actual)
            if acc.id_tipo_cuenta in [1, 2]:  # Activos
                curr_activos += val
            elif acc.id_tipo_cuenta in [3, 4]:  # Pasivos
                curr_pasivos += val
            elif acc.id_tipo_cuenta == 5:  # Capital
                curr_capital += val
        
        curr_ingresos = sum(float(acc.saldo_actual) for acc in income_accounts if acc.id_tipo_cuenta == 6)
        curr_gastos = sum(float(acc.saldo_actual) for acc in income_accounts if acc.id_tipo_cuenta == 7)
        curr_utilidad = curr_ingresos - curr_gastos
        
        # Get specific accounts for current period
        curr_inventarios = get_account_value(balance_accounts, "Inventarios")
        curr_cxc = get_account_value(balance_accounts, "Cuentas por Cobrar")
        
        # PREVIOUS period totals (hardcoded)
        prev_activos = 86000
        prev_pasivos = 0
        prev_capital = 86000
        prev_ingresos = 12800
        prev_gastos = 12800
        prev_utilidad = 0
        prev_inventarios = 0
        prev_cxc = 0
        
        # Safe division helper
        def safe_div(a, b):
            try:
                return (a / b) if b not in (0, None) else None
            except:
                return None
        
        # Calculate all ratios for BOTH periods
        ratios_data = []
        
        # 1. Liquidez Corriente
        ratios_data.append({
            "Ratio": "Liquidez Corriente",
            "Año Base": safe_div(prev_activos, prev_pasivos) if prev_pasivos else None,
            "Año Actual": safe_div(curr_activos, curr_pasivos) if curr_pasivos else None
        })
        
        # 2. Prueba Ácida
        ratios_data.append({
            "Ratio": "Prueba Ácida",
            "Año Base": safe_div(prev_activos - prev_inventarios, prev_pasivos) if prev_pasivos else None,
            "Año Actual": safe_div(curr_activos - curr_inventarios, curr_pasivos) if curr_pasivos else None
        })
        
        # 3. Capital de Trabajo
        ratios_data.append({
            "Ratio": "Capital de Trabajo",
            "Año Base": prev_activos - prev_pasivos,
            "Año Actual": curr_activos - curr_pasivos
        })
        
        # 4. CNO (Capital Neto de Trabajo Operativo)
        # CNO = Activos Corrientes Operativos - Pasivos Corrientes Operativos
        # Para simplificar, usamos: CNO = (Activos - Efectivo) - Pasivos
        prev_efectivo = 0  # From previous period data
        curr_efectivo = get_account_value(balance_accounts, "Efectivo")
        
        ratios_data.append({
            "Ratio": "CNO (Capital Neto de Trabajo Operativo)",
            "Año Base": (prev_activos - prev_efectivo) - prev_pasivos,
            "Año Actual": (curr_activos - curr_efectivo) - curr_pasivos
        })
        
        # 5. Rotación de Inventarios
        ratios_data.append({
            "Ratio": "Rotación de Inventarios",
            "Año Base": safe_div(prev_gastos, prev_inventarios) if prev_inventarios else None,
            "Año Actual": safe_div(curr_gastos, curr_inventarios) if curr_inventarios else None
        })
        
        # 5. Periodo Promedio de Cobro
        ratios_data.append({
            "Ratio": "Periodo Promedio de Cobro (días)",
            "Año Base": (safe_div(prev_cxc, prev_ingresos) * 365) if prev_ingresos and prev_cxc else None,
            "Año Actual": (safe_div(curr_cxc, curr_ingresos) * 365) if curr_ingresos and curr_cxc else None
        })
        
        # 6. Rotación de Activos Totales
        ratios_data.append({
            "Ratio": "Rotación de Activos Totales",
            "Año Base": safe_div(prev_ingresos, prev_activos),
            "Año Actual": safe_div(curr_ingresos, curr_activos)
        })
        
        # 7. Índice de Endeudamiento
        ratios_data.append({
            "Ratio": "Índice de Endeudamiento (Pasivo/Activo)",
            "Año Base": safe_div(prev_pasivos, prev_activos),
            "Año Actual": safe_div(curr_pasivos, curr_activos)
        })
        
        # 8. Margen de Utilidad Bruta
        ratios_data.append({
            "Ratio": "Margen de Utilidad Bruta",
            "Año Base": safe_div(prev_ingresos - prev_gastos, prev_ingresos),
            "Año Actual": safe_div(curr_ingresos - curr_gastos, curr_ingresos)
        })
        
        # 9. Margen de Utilidad Operativa (same as bruta for simplified model)
        ratios_data.append({
            "Ratio": "Margen de Utilidad Operativa",
            "Año Base": safe_div(prev_utilidad, prev_ingresos),
            "Año Actual": safe_div(curr_utilidad, curr_ingresos)
        })
        
        # 10. Margen de Utilidad Neta
        ratios_data.append({
            "Ratio": "Margen de Utilidad Neta",
            "Año Base": safe_div(prev_utilidad, prev_ingresos),
            "Año Actual": safe_div(curr_utilidad, curr_ingresos)
        })
        
        # 11. ROA
        ratios_data.append({
            "Ratio": "ROA (Rent. sobre Activos)",
            "Año Base": safe_div(prev_utilidad, prev_activos),
            "Año Actual": safe_div(curr_utilidad, curr_activos)
        })
        
        # 12. ROE
        ratios_data.append({
            "Ratio": "ROE (Rent. sobre Patrimonio)",
            "Año Base": safe_div(prev_utilidad, prev_capital),
            "Año Actual": safe_div(curr_utilidad, curr_capital)
        })
        
        # 13. Cobertura de Intereses (EBIT/Intereses) - N/A for this model
        ratios_data.append({
            "Ratio": "Cobertura de Intereses (EBIT/Intereses)",
            "Año Base": None,
            "Año Actual": None
        })
        
        # 14. DuPont 3 pasos
        prev_margen = safe_div(prev_utilidad, prev_ingresos)
        prev_rot = safe_div(prev_ingresos, prev_activos)
        prev_mult = safe_div(prev_activos, prev_capital)
        prev_dupont3 = None
        if all(x is not None for x in [prev_margen, prev_rot, prev_mult]):
            prev_dupont3 = prev_margen * prev_rot * prev_mult
        
        curr_margen = safe_div(curr_utilidad, curr_ingresos)
        curr_rot = safe_div(curr_ingresos, curr_activos)
        curr_mult = safe_div(curr_activos, curr_capital)
        curr_dupont3 = None
        if all(x is not None for x in [curr_margen, curr_rot, curr_mult]):
            curr_dupont3 = curr_margen * curr_rot * curr_mult
        
        ratios_data.append({
            "Ratio": "DuPont (3 pasos) - ROE",
            "Año Base": prev_dupont3,
            "Año Actual": curr_dupont3
        })
        
        # 15. DuPont 5 pasos - N/A for simplified model
        ratios_data.append({
            "Ratio": "DuPont (5 pasos) - ROE Extendido",
            "Año Base": None,
            "Año Actual": None
        })
        
        # Calculate change percentage
        for ratio in ratios_data:
            base = ratio["Año Base"]
            curr = ratio["Año Actual"]
            if base is not None and curr is not None and base != 0:
                ratio["Cambio %"] = ((curr - base) / abs(base)) * 100
            else:
                ratio["Cambio %"] = None
        
        self.ratios_data = ratios_data
        df_ratios = pd.DataFrame(ratios_data)
        
        # Render UI
        tab = self.tabs["Razones"]
        for widget in tab.winfo_children():
            widget.destroy()
        
        sf = ctk.CTkScrollableFrame(tab)
        sf.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Header
        header = ctk.CTkFrame(sf, fg_color="transparent")
        header.pack(fill="x", pady=(0, 8))
        ctk.CTkLabel(header, text="Razón", font=ctk.CTkFont(weight="bold"), width=300, anchor="w").pack(side="left")
        ctk.CTkLabel(header, text="Año Base", font=ctk.CTkFont(weight="bold"), width=150).pack(side="left")
        ctk.CTkLabel(header, text="Año Actual", font=ctk.CTkFont(weight="bold"), width=150).pack(side="left")
        ctk.CTkLabel(header, text="Cambio %", font=ctk.CTkFont(weight="bold"), width=120).pack(side="left")
        
        # Rows
        for _, row in df_ratios.iterrows():
            name = row["Ratio"]
            base = row["Año Base"]
            curr = row["Año Actual"]
            change = row["Cambio %"]
            
            rframe = ctk.CTkFrame(sf, fg_color=get_color('card_bg'))
            rframe.pack(fill="x", pady=4, padx=2)
            
            # Ratio name
            ctk.CTkLabel(rframe, text=name, width=300, anchor="w", font=ctk.CTkFont(weight="bold"), text_color="#1f2937").pack(side="left", padx=(6, 0))
            
            # Format values
            def format_ratio(val, ratio_name):
                if val is None:
                    return "-"
                # Check for NaN or infinity
                try:
                    import math
                    if math.isnan(val) or math.isinf(val):
                        return "-"
                except:
                    pass
                # Show as percentage for certain ratios
                if any(k in ratio_name.lower() for k in ['margen', 'roa', 'roe', 'endeudamiento', 'dupont']):
                    return f"{val * 100:.2f}%"
                # Show as days
                if 'días' in ratio_name.lower():
                    return f"{val:.0f} días"
                # Show as currency for Capital de Trabajo and CNO
                if 'capital de trabajo' in ratio_name.lower() or 'cno' in ratio_name.lower():
                    return f"${val:,.2f}"
                return f"{val:.2f}"

            
            ctk.CTkLabel(rframe, text=format_ratio(base, name), width=150, anchor="center").pack(side="left")
            ctk.CTkLabel(rframe, text=format_ratio(curr, name), width=150, anchor="center").pack(side="left")
            
            # Change % with color
            if change is None:
                change_text = "-"
                color = "#374151"
            else:
                sign = '+' if change >= 0 else ''
                change_text = f"{sign}{change:.2f}%"
                color = "#16a34a" if change >= 0 else "#dc2626"
            
            ctk.CTkLabel(rframe, text=change_text, width=120, anchor="center", text_color=color, font=ctk.CTkFont(weight="bold")).pack(side="left")

    def load_cash_flow_statements(self):
        """Display cash flow statements using both direct and indirect methods"""
        tab = self.tabs["Flujos"]
        for widget in tab.winfo_children():
            widget.destroy()
        
        # Fetch data
        income_accounts = CuentaDAO.obtener_cuentas_resultados_plataforma()
        balance_accounts = CuentaDAO.obtener_cuentas_balance_plataforma()
        
        # Helper to normalize names and get account values
        import unicodedata
        def normalize(name):
            normalized = unicodedata.normalize('NFD', name)
            normalized = ''.join(c for c in normalized if unicodedata.category(c) != 'Mn')
            return normalized.lower().strip()
        
        def get_account_value(accounts_list, account_name):
            target_normalized = normalize(account_name)
            for acc in accounts_list:
                if normalize(acc.nombre_cuenta) == target_normalized:
                    return float(acc.saldo_actual)
            return 0.0
        
        # Get specific accounts from database
        ingresos = get_account_value(income_accounts, "Ingresos")
        gasto_envio = get_account_value(income_accounts, "Gasto de envio")
        renta = get_account_value(income_accounts, "Renta")
        
        efectivo_actual = get_account_value(balance_accounts, "Efectivo")
        mobiliario = get_account_value(balance_accounts, "Mobiliario y accesorios")
        vehiculos = get_account_value(balance_accounts, "Vehiculos")
        dep_mobiliario = get_account_value(balance_accounts, "Depreciacion de mobiliario y accesorios")
        dep_vehiculos = get_account_value(balance_accounts, "Depreciacion de vehiculos")
        capital_social = get_account_value(balance_accounts, "Capital social")
        utilidades_retenidas = get_account_value(balance_accounts, "Utilidades retenidas")
        
        # Calculate totals
        total_gastos = gasto_envio + renta
        utilidad_neta = ingresos - total_gastos
        
        # Previous period values (hardcoded as per original data)
        efectivo_inicial = 0
        mobiliario_inicial = 36000
        vehiculos_inicial = 50000
        
        # Calculate changes
        compra_mobiliario = max(0, mobiliario - mobiliario_inicial)
        compra_vehiculos = max(0, vehiculos - vehiculos_inicial)
        total_compra_activos = compra_mobiliario + compra_vehiculos
        
        aportacion_capital = max(0, capital_social - 86000)  # Initial capital was 86000
        
        # Total depreciation
        depreciacion_total = dep_mobiliario + dep_vehiculos
        
        # Container for side-by-side display
        container = ctk.CTkFrame(tab, fg_color="transparent")
        container.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Left side: Direct Method
        left_frame = ctk.CTkFrame(container, fg_color="transparent")
        left_frame.pack(side='left', fill='both', expand=True, padx=(0, 5))
        
        ctk.CTkLabel(left_frame, text="Método Directo", font=ctk.CTkFont(size=16, weight='bold'), text_color="#111827").pack(anchor='w', pady=(0, 10))
        
        direct_scroll = ctk.CTkScrollableFrame(left_frame)
        direct_scroll.pack(fill='both', expand=True)
        
        # Calculate operating cash flow (direct method)
        cobros_clientes = ingresos  # Assuming all sales are cash
        pagos_operativos = -(gasto_envio + renta)  # Negative because they are outflows
        efectivo_operacion_directo = cobros_clientes + pagos_operativos
        
        # Calculate net increase in cash (sum of all three activities)
        aumento_neto_efectivo = efectivo_operacion_directo - total_compra_activos + aportacion_capital
        
        # Direct Method Data
        direct_data = [
            ("ACTIVIDADES DE OPERACIÓN", None, True),
            ("Cobros de clientes (Ingresos)", cobros_clientes, False),
            ("Pago de gasto de envío", -gasto_envio, False),
            ("Pago de renta", -renta, False),
            ("Efectivo neto de actividades de operación", efectivo_operacion_directo, True),
            ("", None, False),
            ("ACTIVIDADES DE INVERSIÓN", None, True),
            ("Compra de mobiliario y accesorios", -compra_mobiliario if compra_mobiliario > 0 else 0, False),
            ("Compra de vehículos", -compra_vehiculos if compra_vehiculos > 0 else 0, False),
            ("Efectivo neto de actividades de inversión", -total_compra_activos, True),
            ("", None, False),
            ("ACTIVIDADES DE FINANCIAMIENTO", None, True),
            ("Aportaciones de capital", aportacion_capital, False),
            ("Pago de dividendos", 0, False),
            ("Efectivo neto de actividades de financiamiento", aportacion_capital, True),
            ("", None, False),
            ("Aumento (disminución) neto en efectivo", aumento_neto_efectivo, True),
            ("Efectivo al inicio del período", efectivo_inicial, False),
            ("Efectivo al final del período", efectivo_inicial + aumento_neto_efectivo, True)
        ]
        
        for label, value, is_bold in direct_data:
            if value is None:
                # Header
                ctk.CTkLabel(direct_scroll, text=label, font=ctk.CTkFont(weight='bold', size=13), text_color="#1f2937").pack(anchor='w', pady=(8, 2))
            elif label == "":
                # Spacer
                ctk.CTkFrame(direct_scroll, height=5, fg_color="transparent").pack()
            else:
                row = ctk.CTkFrame(direct_scroll, fg_color="transparent")
                row.pack(fill='x', pady=2)
                
                font = ctk.CTkFont(weight='bold' if is_bold else 'normal')
                color = "#111827" if is_bold else "#374151"
                
                ctk.CTkLabel(row, text=label, font=font, text_color=color, anchor='w').pack(side='left', fill='x', expand=True)
                ctk.CTkLabel(row, text=f"${value:,.2f}", font=font, text_color=color, anchor='e').pack(side='right')
        
        # Right side: Indirect Method
        right_frame = ctk.CTkFrame(container, fg_color="transparent")
        right_frame.pack(side='left', fill='both', expand=True, padx=(5, 0))
        
        ctk.CTkLabel(right_frame, text="Método Indirecto", font=ctk.CTkFont(size=16, weight='bold'), text_color="#111827").pack(anchor='w', pady=(0, 10))
        
        indirect_scroll = ctk.CTkScrollableFrame(right_frame)
        indirect_scroll.pack(fill='both', expand=True)
        
        # Indirect Method Data
        # Starting with net income and adjusting for non-cash items
        efectivo_operacion_indirecto = utilidad_neta + depreciacion_total
        
        # Calculate net increase in cash (same as direct method - must match!)
        aumento_neto_efectivo_indirecto = efectivo_operacion_indirecto - total_compra_activos + aportacion_capital
        
        indirect_data = [
            ("ACTIVIDADES DE OPERACIÓN", None, True),
            ("Utilidad neta", utilidad_neta, False),
            ("Ajustes para conciliar utilidad neta:", None, False),
            ("  Depreciación de mobiliario", dep_mobiliario, False),
            ("  Depreciación de vehículos", dep_vehiculos, False),
            ("Efectivo neto de actividades de operación", efectivo_operacion_indirecto, True),
            ("", None, False),
            ("ACTIVIDADES DE INVERSIÓN", None, True),
            ("Compra de mobiliario y accesorios", -compra_mobiliario if compra_mobiliario > 0 else 0, False),
            ("Compra de vehículos", -compra_vehiculos if compra_vehiculos > 0 else 0, False),
            ("Efectivo neto de actividades de inversión", -total_compra_activos, True),
            ("", None, False),
            ("ACTIVIDADES DE FINANCIAMIENTO", None, True),
            ("Aportaciones de capital", aportacion_capital, False),
            ("Pago de dividendos", 0, False),
            ("Efectivo neto de actividades de financiamiento", aportacion_capital, True),
            ("", None, False),
            ("Aumento (disminución) neto en efectivo", aumento_neto_efectivo_indirecto, True),
            ("Efectivo al inicio del período", efectivo_inicial, False),
            ("Efectivo al final del período", efectivo_inicial + aumento_neto_efectivo_indirecto, True)
        ]
        
        for label, value, is_bold in indirect_data:
            if value is None and label and not label.startswith("  "):
                # Header
                ctk.CTkLabel(indirect_scroll, text=label, font=ctk.CTkFont(weight='bold', size=13), text_color="#1f2937").pack(anchor='w', pady=(8, 2))
            elif label == "":
                # Spacer
                ctk.CTkFrame(indirect_scroll, height=5, fg_color="transparent").pack()
            elif value is None:
                # Sub-header (no value)
                ctk.CTkLabel(indirect_scroll, text=label, font=ctk.CTkFont(size=11, slant='italic'), text_color="#6b7280").pack(anchor='w', pady=1)
            else:
                row = ctk.CTkFrame(indirect_scroll, fg_color="transparent")
                row.pack(fill='x', pady=2)
                
                font = ctk.CTkFont(weight='bold' if is_bold else 'normal')
                color = "#111827" if is_bold else "#374151"
                
                ctk.CTkLabel(row, text=label, font=font, text_color=color, anchor='w').pack(side='left', fill='x', expand=True)
                ctk.CTkLabel(row, text=f"${value:,.2f}", font=font, text_color=color, anchor='e').pack(side='right')

    def load_charts(self):
        """Display charts: pie charts for asset/liability breakdown and bar chart for comparison"""
        # Fetch current period data
        balance_accounts = CuentaDAO.obtener_cuentas_balance_plataforma()
        
        # Helper to normalize names
        import unicodedata
        def normalize_name(name):
            normalized = unicodedata.normalize('NFD', name)
            normalized = ''.join(c for c in normalized if unicodedata.category(c) != 'Mn')
            return normalized.lower().strip()
        
        # Organize current period accounts
        activos_data = {}
        pasivos_data = {}
        capital_data = {}
        
        for acc in balance_accounts:
            name = acc.nombre_cuenta
            val = float(acc.saldo_actual)
            
            if acc.id_tipo_cuenta in [1, 2]:  # Activos
                activos_data[name] = val
            elif acc.id_tipo_cuenta in [3, 4]:  # Pasivos
                pasivos_data[name] = val
            elif acc.id_tipo_cuenta == 5:  # Capital
                capital_data[name] = val
        
        # Calculate totals
        total_activos_curr = sum(activos_data.values())
        total_pasivos_curr = sum(pasivos_data.values())
        total_capital_curr = sum(capital_data.values())
        
        # Previous period totals
        total_activos_prev = 86000
        
        # Clear tab
        tab = self.tabs["Graficos"]
        for widget in tab.winfo_children():
            widget.destroy()
        
        # Create scrollable frame
        scroll_frame = ctk.CTkScrollableFrame(tab)
        scroll_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Top container for pie charts (side by side)
        top_frame = ctk.CTkFrame(scroll_frame, fg_color="transparent")
        top_frame.pack(fill="x", pady=(0, 20))
        
        # === PIE CHART 1: ACTIVOS ===
        fig1, ax1 = plt.subplots(figsize=(5, 5.5))  # Increased height for bottom legend
        
        if activos_data:
            # Filter out very small values and group them as "Otros"
            threshold = total_activos_curr * 0.03  # 3% threshold
            main_items = {}
            otros_total = 0
            
            for name, val in activos_data.items():
                if val >= threshold:
                    main_items[name] = val
                else:
                    otros_total += val
            
            if otros_total > 0:
                main_items['Otros'] = otros_total
            
            labels1 = list(main_items.keys())
            sizes1 = list(main_items.values())
            colors1 = plt.cm.Set3(range(len(labels1)))
            
            # Use autopct only, no labels on the pie itself
            wedges, texts, autotexts = ax1.pie(sizes1, autopct='%1.1f%%', startangle=90, colors=colors1)
            
            # Make percentage text bold and BLACK
            for autotext in autotexts:
                autotext.set_color('black')
                autotext.set_fontweight('bold')
                autotext.set_fontsize(10)
            
            # Add legend BELOW the pie
            ax1.legend(wedges, labels1, title="Cuentas", loc="upper center", bbox_to_anchor=(0.5, -0.05), fontsize=9, ncol=2)
            ax1.set_title('Composición de Activos\nPeríodo Actual', fontsize=13, fontweight='bold', pad=10)
        else:
            ax1.text(0.5, 0.5, 'No hay datos de activos', ha='center', va='center')
            ax1.set_title('Composición de Activos\nPeríodo Actual', fontsize=13, fontweight='bold', pad=10)
        
        plt.tight_layout()
        canvas1 = FigureCanvasTkAgg(fig1, master=top_frame)
        canvas1.draw()
        canvas1.get_tk_widget().pack(side="left", padx=10)
        
        # === PIE CHART 2: PASIVOS + CAPITAL ===
        fig2, ax2 = plt.subplots(figsize=(5, 5.5))  # Increased height for bottom legend
        
        # Combine pasivos and capital
        pasivo_capital_data = {}
        pasivo_capital_data.update(pasivos_data)
        pasivo_capital_data.update(capital_data)
        
        total_pasivo_capital = total_pasivos_curr + total_capital_curr
        
        if pasivo_capital_data:
            # Filter small values
            threshold2 = total_pasivo_capital * 0.03 if total_pasivo_capital > 0 else 0
            main_items2 = {}
            otros_total2 = 0
            
            for name, val in pasivo_capital_data.items():
                if val >= threshold2:
                    main_items2[name] = val
                else:
                    otros_total2 += val
            
            if otros_total2 > 0:
                main_items2['Otros'] = otros_total2
            
            labels2 = list(main_items2.keys())
            sizes2 = list(main_items2.values())
            colors2 = plt.cm.Set2(range(len(labels2)))
            
            wedges2, texts2, autotexts2 = ax2.pie(sizes2, autopct='%1.1f%%', startangle=90, colors=colors2)
            
            for autotext in autotexts2:
                autotext.set_color('black')
                autotext.set_fontweight('bold')
                autotext.set_fontsize(10)
            
            ax2.legend(wedges2, labels2, title="Cuentas", loc="upper center", bbox_to_anchor=(0.5, -0.05), fontsize=9, ncol=2)
            ax2.set_title('Composición de Pasivos + Capital\nPeríodo Actual', fontsize=13, fontweight='bold', pad=10)
        else:
            ax2.text(0.5, 0.5, 'No hay datos de pasivos/capital', ha='center', va='center')
            ax2.set_title('Composición de Pasivos + Capital\nPeríodo Actual', fontsize=13, fontweight='bold', pad=10)
        
        plt.tight_layout()
        canvas2 = FigureCanvasTkAgg(fig2, master=top_frame)
        canvas2.draw()
        canvas2.get_tk_widget().pack(side="left", padx=10)
        
        # === BAR CHART: COMPARISON ===
        fig3, ax3 = plt.subplots(figsize=(8, 5))
        
        periods = ['Período Anterior', 'Período Actual']
        values = [total_activos_prev, total_activos_curr]
        colors3 = ['#E5E7EB', '#06B6D4']
        
        bars = ax3.bar(periods, values, color=colors3, width=0.6)
        ax3.set_title('Comparación de Activos Totales', fontsize=14, fontweight='bold')
        ax3.set_ylabel('Monto (C$)', fontsize=12)
        
        # Add value labels on bars
        for bar in bars:
            height = bar.get_height()
            ax3.text(bar.get_x() + bar.get_width()/2., height,
                    f'${height:,.0f}',
                    ha='center', va='bottom', fontsize=11, fontweight='bold')
        
        ax3.grid(axis='y', alpha=0.3)
        
        canvas3 = FigureCanvasTkAgg(fig3, master=scroll_frame)
        canvas3.draw()
        canvas3.get_tk_widget().pack(pady=10)

    def render_table(self, parent, df):
        if df is None or df.empty:
            ctk.CTkLabel(parent, text="No hay datos disponibles").pack()
            return

        # Create frame for table
        table_frame = ctk.CTkFrame(parent, fg_color="transparent")
        table_frame.pack(fill="x", expand=True)

        # Headers
        headers = list(df.columns)
        header_frame = ctk.CTkFrame(table_frame, fg_color=get_color('card_bg'))
        header_frame.pack(fill="x", pady=(0, 2))
        
        cols = len(headers)
        for i, h in enumerate(headers):
            ctk.CTkLabel(header_frame, text=str(h), font=ctk.CTkFont(weight="bold"), width=150, anchor="w").pack(side="left", padx=5)

        # Rows
        for _, row in df.iterrows():
            row_frame = ctk.CTkFrame(table_frame, fg_color="transparent")
            row_frame.pack(fill="x", pady=1)
            
            for col in headers:
                val = row[col]
                # Format value
                if isinstance(val, (int, float)):
                    if "Vertical" in col or "%" in col:
                        text_val = f"{val:.2f}%"
                    else:
                        text_val = f"${val:,.2f}"
                else:
                    text_val = str(val)
                
                ctk.CTkLabel(row_frame, text=text_val, width=150, anchor="w").pack(side="left", padx=5)

    def render_proforma(self, parent):
        for w in parent.winfo_children():
            w.destroy()

        top = ctk.CTkFrame(parent, fg_color="transparent")
        top.pack(fill='x', padx=8, pady=6)
        ctk.CTkLabel(top, text="Proforma - Proyección", font=ctk.CTkFont(size=16, weight='bold'), text_color="#111827").pack(side='left')

        # Switch for Mode
        self.proforma_mode = ctk.StringVar(value="pct") # 'pct' or 'amount'
        
        def toggle_mode():
            mode = self.proforma_mode.get()
            # Update labels
            for key, label_widget in self.field_labels.items():
                base_text = self.field_base_names[key]
                if mode == "amount":
                    label_widget.configure(text=f"Nuevo Monto {base_text} ($)")
                else:
                    label_widget.configure(text=f"Crecimiento {base_text} (%)")
        
        switch = ctk.CTkSwitch(top, text="Modo: Monto ($)", variable=self.proforma_mode, onvalue="amount", offvalue="pct", command=toggle_mode)
        switch.pack(side='right')
        ctk.CTkLabel(top, text="Modo: Porcentaje (%)", text_color="#374151").pack(side='right', padx=10)

        form = ctk.CTkFrame(parent)
        form.pack(fill='x', padx=8, pady=6)

        # Campos específicos
        self.pct_entries = {}
        self.field_labels = {}
        self.field_base_names = {
            'ingresos': 'Ingresos',
            'envio': 'Gasto de Envío',
            'renta': 'Renta'
        }
        
        fields_order = ['ingresos', 'envio', 'renta']

        for key in fields_order:
            row = ctk.CTkFrame(form, fg_color='transparent')
            row.pack(fill='x', pady=2)
            
            base_name = self.field_base_names[key]
            lbl_text = f"Crecimiento {base_name} (%)"
            
            lbl = ctk.CTkLabel(row, text=lbl_text, width=250, anchor='w', text_color="#374151")
            lbl.pack(side='left', padx=(0,6))
            self.field_labels[key] = lbl
            
            ent = ctk.CTkEntry(row, width=120)
            ent.pack(side='left')
            self.pct_entries[key] = ent

        btn_row = ctk.CTkFrame(parent, fg_color='transparent')
        btn_row.pack(fill='x', padx=8, pady=6)
        ctk.CTkButton(btn_row, text='Generar Proforma', command=self.apply_proforma, fg_color='#06B6D4').pack(side='left')

        # Result area
        self.proforma_result = ctk.CTkScrollableFrame(parent)
        self.proforma_result.pack(fill='both', expand=True, padx=8, pady=6)

    def apply_proforma(self):
        mode = self.proforma_mode.get()
        
        # Read inputs
        inputs = {}
        for key, ent in self.pct_entries.items():
            v = ent.get().strip()
            if v == '':
                inputs[key] = 0.0
            else:
                try:
                    inputs[key] = float(v)
                except:
                    inputs[key] = 0.0

        # Fetch current data
        income_accounts = CuentaDAO.obtener_cuentas_resultados_plataforma()
        balance_accounts = CuentaDAO.obtener_cuentas_balance_plataforma()
        
        # Helper to normalize names
        import unicodedata
        def normalize(name):
            normalized = unicodedata.normalize('NFD', name)
            normalized = ''.join(c for c in normalized if unicodedata.category(c) != 'Mn')
            return normalized.lower().strip()

        # --- 1. Proyectar Estado de Resultados ---
        is_data = []
        total_ingresos_proj = 0
        total_gastos_proj = 0
        
        for acc in income_accounts:
            name = acc.nombre_cuenta
            val = float(acc.saldo_actual)
            norm_name = normalize(name)
            
            new_val = val
            
            if acc.id_tipo_cuenta == 6: # Ingresos
                # Apply Ingresos factor
                if mode == "pct":
                    new_val = val * (1.0 + inputs['ingresos'] / 100.0)
                else:
                    # Target amount logic: distribute target proportionally if multiple income accounts
                    # For simplicity, if multiple, we scale all by (Target / Total_Curr)
                    # But here we iterate. Let's calculate total current revenue first.
                    pass 
            
            is_data.append({"Cuenta": name, "Monto": val, "TipoId": acc.id_tipo_cuenta, "NormName": norm_name})

        # Calculate current totals for scaling
        curr_total_ingresos = sum(d["Monto"] for d in is_data if d["TipoId"] == 6)
        curr_envio = sum(d["Monto"] for d in is_data if 'envio' in d["NormName"])
        curr_renta = sum(d["Monto"] for d in is_data if 'renta' in d["NormName"] or 'alquiler' in d["NormName"])
        
        # Apply projections
        final_is_data = []
        
        for item in is_data:
            val = item["Monto"]
            tipo_id = item["TipoId"]
            norm_name = item["NormName"]
            new_val = val
            
            if tipo_id == 6: # Ingresos
                if mode == "pct":
                    new_val = val * (1.0 + inputs['ingresos'] / 100.0)
                else:
                    # Target amount
                    if curr_total_ingresos > 0:
                        ratio = val / curr_total_ingresos
                        new_val = inputs['ingresos'] * ratio if inputs['ingresos'] > 0 else val # If input 0/empty, keep same? Or 0? Assuming input is target.
                        if inputs['ingresos'] == 0 and self.pct_entries['ingresos'].get().strip() == "":
                             new_val = val # Empty input -> no change
                    else:
                        new_val = inputs['ingresos'] # If 0 current, set to target (if single account)
                
                total_ingresos_proj += new_val
                
            elif tipo_id == 7: # Gastos
                # Check for specific drivers
                if 'envio' in norm_name:
                    if mode == "pct":
                        new_val = val * (1.0 + inputs['envio'] / 100.0)
                    else:
                        # Target amount
                        if curr_envio > 0:
                            ratio = val / curr_envio
                            target = inputs['envio']
                            if target == 0 and self.pct_entries['envio'].get().strip() == "":
                                target = curr_envio
                            new_val = target * ratio
                        else:
                            new_val = inputs['envio']
                            
                elif 'renta' in norm_name or 'alquiler' in norm_name:
                    if mode == "pct":
                        new_val = val * (1.0 + inputs['renta'] / 100.0)
                    else:
                        # Target amount
                        if curr_renta > 0:
                            ratio = val / curr_renta
                            target = inputs['renta']
                            if target == 0 and self.pct_entries['renta'].get().strip() == "":
                                target = curr_renta
                            new_val = target * ratio
                        else:
                            new_val = inputs['renta']
                
                # Else keep constant
                total_gastos_proj += new_val
            
            final_is_data.append({"Cuenta": item["Cuenta"], "Monto": new_val})

        # Calculate IR and Net Income
        utilidad_antes_impuestos = total_ingresos_proj - total_gastos_proj
        impuestos = max(0, utilidad_antes_impuestos * 0.30) # 30% IR, only if profit
        utilidad_neta_proj = utilidad_antes_impuestos - impuestos
        
        # Add IR and Net Income to IS display
        final_is_data.append({"Cuenta": "Utilidad Antes de Impuestos", "Monto": utilidad_antes_impuestos})
        final_is_data.append({"Cuenta": "Impuestos (30%)", "Monto": impuestos})
        final_is_data.append({"Cuenta": "Utilidad Neta", "Monto": utilidad_neta_proj})
        
        # Vertical Analysis IS
        for row in final_is_data:
            row["Vertical %"] = (row["Monto"] / total_ingresos_proj * 100) if total_ingresos_proj else 0.0
        
        df_is_proj = pd.DataFrame(final_is_data)
        self.proforma_is_df = df_is_proj # Store for export

        # --- 2. Proyectar Balance General ---
        # Logic:
        # Assets: Efectivo += Utilidad Neta (After Tax)
        # Liabilities: Constant
        # Equity: Utilidades Retenidas += Utilidad Neta (After Tax)
        
        bs_data = []
        total_activos_proj = 0
        
        for acc in balance_accounts:
            name = acc.nombre_cuenta
            val = float(acc.saldo_actual)
            tipo_id = acc.id_tipo_cuenta
            norm_name = normalize(name)
            
            new_val = val
            
            if 'efectivo' in norm_name or 'banco' in norm_name or 'caja' in norm_name:
                # Add Net Income to Cash
                new_val = val + utilidad_neta_proj
            
            if 'utilidades retenidas' in norm_name or 'resultados acumulados' in norm_name:
                # Add Net Income to Retained Earnings
                new_val = val + utilidad_neta_proj
                
            bs_data.append({"Cuenta": name, "Monto": new_val, "TipoId": tipo_id})
            
        # Calculate totals
        total_activos_proj = sum(d["Monto"] for d in bs_data if d["TipoId"] in [1, 2])
        
        # Vertical Analysis BS
        for row in bs_data:
            row["Vertical %"] = (row["Monto"] / total_activos_proj * 100) if total_activos_proj else 0.0
        
        df_bs_proj = pd.DataFrame(bs_data)
        if not df_bs_proj.empty:
            df_bs_proj = df_bs_proj.drop(columns=["TipoId"])
            
        self.proforma_bs_df = df_bs_proj # Store for export

        # --- Render Results ---
        for w in self.proforma_result.winfo_children():
            w.destroy()

        # Container for side-by-side tables
        tables_container = ctk.CTkFrame(self.proforma_result, fg_color="transparent")
        tables_container.pack(fill='x', expand=True, pady=10)
        
        # Left Frame: Balance General
        left_frame = ctk.CTkFrame(tables_container, fg_color="transparent")
        left_frame.pack(side='left', fill='both', expand=True, padx=(0, 10))
        
        ctk.CTkLabel(left_frame, text='Balance Proforma (Proyectado)', font=ctk.CTkFont(weight='bold'), text_color='#111827').pack(anchor='w', pady=(0,4))
        self.render_table(left_frame, df_bs_proj)

        # Right Frame: Estado de Resultados
        right_frame = ctk.CTkFrame(tables_container, fg_color="transparent")
        right_frame.pack(side='left', fill='both', expand=True, padx=(10, 0))
        
        ctk.CTkLabel(right_frame, text='Estado de Resultados Proforma (Proyectado)', font=ctk.CTkFont(weight='bold'), text_color='#111827').pack(anchor='w', pady=(0,4))
        self.render_table(right_frame, df_is_proj)
        
        # Show Net Income Summary (Bottom)
        summary_frame = ctk.CTkFrame(self.proforma_result, fg_color="#F3F4F6")
        summary_frame.pack(fill='x', pady=10, padx=5)
        ctk.CTkLabel(summary_frame, text=f"Utilidad Antes de Impuestos: ${utilidad_antes_impuestos:,.2f}", font=ctk.CTkFont(size=12), text_color="#374151").pack(pady=(5,0))
        ctk.CTkLabel(summary_frame, text=f"Impuestos (30%): ${impuestos:,.2f}", font=ctk.CTkFont(size=12), text_color="#dc2626").pack(pady=(0,0))
        ctk.CTkLabel(summary_frame, text=f"Utilidad Neta Proyectada: ${utilidad_neta_proj:,.2f}", font=ctk.CTkFont(weight="bold", size=14), text_color="#1f2937").pack(pady=5)
        ctk.CTkLabel(summary_frame, text="Nota: La utilidad neta (después de impuestos) se ha sumado a 'Efectivo' y 'Utilidades Retenidas'.", font=ctk.CTkFont(size=11), text_color="#6B7280").pack(pady=(0,5))

    def show_interpretations(self):
        # Generate simple textual interpretations based on ratios and trends
        text = []
        
        if not hasattr(self, 'ratios_data') or not self.ratios_data:
            text.append('No hay datos suficientes para generar interpretaciones.')
        else:
            # Helper to find ratio value
            def get_val(name_part):
                r = next((r for r in self.ratios_data if name_part in r["Ratio"]), None)
                return r["Año Actual"] if r else None

            # Liquidity
            val = get_val("Liquidez Corriente")
            if val is not None:
                if val == 0:
                    text.append('Liquidez Corriente es 0: No hay Activos Corrientes.')
                elif val < 1:
                    text.append('La liquidez corriente es baja (<1): riesgo de falta de recursos para obligaciones de corto plazo.')
                elif val < 1.5:
                    text.append('La liquidez es moderada; conviene vigilar el capital de trabajo.')
                else:
                    text.append('La liquidez parece adecuada.')
            else:
                text.append('Liquidez Corriente: No calculable (Pasivos Corrientes es 0).')

            # Profitability
            val = get_val("Margen de Utilidad Neta")
            if val is not None:
                if val == 0:
                    text.append('Margen Neto es 0%: La Utilidad Neta es 0.')
                elif val < 0.05:
                    text.append('Margen Neto bajo: revisar estructura de costos y precios.')
                else:
                    text.append('Margen Neto aceptable.')
            else:
                text.append('Margen Neto: No calculable.')

            # Leverage
            val = get_val("Índice de Endeudamiento")
            if val is not None:
                if val == 0:
                    text.append('Nivel de endeudamiento es 0%: No hay pasivos reportados.')
                elif val > 0.6:
                    text.append('Alto apalancamiento: considerar reducir deuda o aumentar patrimonio.')
                else:
                    text.append('Nivel de endeudamiento moderado.')
            else:
                text.append('Endeudamiento: No calculable.')

            text.append('\nRecomendaciones generales:')
            text.append('- Mejorar rotación de activos si las ventas son bajas respecto a activos.')
            text.append('- Revisar inventarios y cuentas por cobrar para liberar capital de trabajo.')

        # Show in a popup
        popup = ctk.CTkToplevel(self)
        popup.geometry('600x400')
        popup.title('Interpretaciones y Recomendaciones')
        body = ctk.CTkScrollableFrame(popup)
        body.pack(fill='both', expand=True, padx=8, pady=8)
        for p in text:
            ctk.CTkLabel(body, text=p, wraplength=560, text_color=get_color('text')).pack(anchor='w', pady=6)

    def load_income_statement(self):
        # 1. Datos del Período Anterior (Hardcoded)
        prev_data_raw = [
            {"Cuenta": "Ingresos", "Monto": 12800},
            {"Cuenta": "Gasto de envío", "Monto": 2800},
            {"Cuenta": "Renta", "Monto": 10000},
            {"Cuenta": "Utilidad Antes de IR", "Monto": 3000},
            {"Cuenta": "IR", "Monto": 0},
            {"Cuenta": "Utilidad del período", "Monto": 0}
        ]
        
        # Calcular Vertical % para Anterior
        prev_ingresos = 12800
        for row in prev_data_raw:
            row["Vertical %"] = (row["Monto"] / prev_ingresos * 100) if prev_ingresos else 0.0
            
        df_prev = pd.DataFrame(prev_data_raw)
        
        # 2. Datos del Período Actual (Desde BD)
        current_accounts = CuentaDAO.obtener_cuentas_resultados_plataforma()
        
        # Agrupar y calcular totales para el período actual
        curr_ingresos = 0
        curr_gastos = 0
        curr_data_map = {}
        
        for acc in current_accounts:
            name = acc.nombre_cuenta
            val = float(acc.saldo_actual)
            curr_data_map[name] = val
            
            if acc.id_tipo_cuenta == 6: # Ingresos
                curr_ingresos += val
            elif acc.id_tipo_cuenta == 7: # Gastos
                curr_gastos += val

        # Calcular utilidades actuales
        curr_util_antes_ir = curr_ingresos - curr_gastos
        curr_ir = 0 
        curr_util_neta = curr_util_antes_ir - curr_ir

        # Agregar totales calculados
        curr_data_map["Ingresos"] = curr_ingresos
        if "Gasto de envio" in curr_data_map:
            curr_data_map["Gasto de envío"] = curr_data_map.pop("Gasto de envio")
        
        curr_data_map["Utilidad Antes de IR"] = curr_util_antes_ir
        curr_data_map["IR"] = curr_ir
        curr_data_map["Utilidad del período"] = curr_util_neta

        # Construir lista para DataFrame Actual (ordenado similar al anterior si es posible)
        curr_data_list = []
        processed_keys = set()
        
        # Primero las cuentas que coinciden con el orden del anterior
        for row in prev_data_raw:
            key = row["Cuenta"]
            if key in curr_data_map:
                val = curr_data_map[key]
                vert = (val / curr_ingresos * 100) if curr_ingresos else 0.0
                curr_data_list.append({"Cuenta": key, "Monto": val, "Vertical %": vert})
                processed_keys.add(key)
        
        # Luego el resto de cuentas
        for key, val in curr_data_map.items():
            if key not in processed_keys:
                vert = (val / curr_ingresos * 100) if curr_ingresos else 0.0
                curr_data_list.append({"Cuenta": key, "Monto": val, "Vertical %": vert})

        df_curr = pd.DataFrame(curr_data_list)
        
        # 3. Renderizar UI (Dos tablas lado a lado)
        tab = self.tabs["Resultados"]
        for widget in tab.winfo_children():
            widget.destroy()

        # Contenedor principal
        container = ctk.CTkFrame(tab, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Frame Izquierdo (Anterior)
        left_frame = ctk.CTkFrame(container, fg_color="transparent")
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        ctk.CTkLabel(left_frame, text="Período Anterior", font=ctk.CTkFont(size=16, weight='bold'), text_color="#111827").pack(anchor='w', pady=(0, 5))
        self.render_table(left_frame, df_prev)
        
        # Frame Derecho (Actual)
        right_frame = ctk.CTkFrame(container, fg_color="transparent")
        right_frame.pack(side="left", fill="both", expand=True, padx=(10, 0))
        
        ctk.CTkLabel(right_frame, text="Período Actual", font=ctk.CTkFont(size=16, weight='bold'), text_color="#111827").pack(anchor='w', pady=(0, 5))
        self.render_table(right_frame, df_curr)

        # 4. Análisis Horizontal (Abajo)
        horizontal_data = []
        
        # Crear un mapa del anterior para búsqueda rápida
        prev_map = {row["Cuenta"]: row["Monto"] for row in prev_data_raw}
        
        # Unir todas las claves
        all_keys = sorted(list(set(prev_map.keys()) | set(curr_data_map.keys())))
        
        # Ordenar: Ingresos primero, luego gastos, luego utilidades
        # (Una heurística simple o usar el orden de prev_data_raw + nuevos)
        ordered_keys = []
        # Primero los que estaban en prev (mantiene orden lógico)
        for row in prev_data_raw:
            if row["Cuenta"] in all_keys:
                ordered_keys.append(row["Cuenta"])
        # Luego los nuevos
        for k in all_keys:
            if k not in ordered_keys:
                ordered_keys.append(k)

        for key in ordered_keys:
            val_prev = prev_map.get(key, 0.0)
            val_curr = curr_data_map.get(key, 0.0)
            var_abs = val_curr - val_prev
            var_pct = (var_abs / val_prev * 100) if val_prev != 0 else 0.0
            
            horizontal_data.append({
                "Cuenta": key,
                "Año Base": val_prev,
                "Año Actual": val_curr,
                "Variación $": var_abs,
                "Variación %": var_pct
            })
            
        df_horizontal = pd.DataFrame(horizontal_data)
        
        ctk.CTkLabel(tab, text="Análisis Horizontal (Variación)", font=ctk.CTkFont(size=16, weight='bold'), text_color="#111827").pack(anchor='w', padx=10, pady=(20, 5))
        self.render_table(tab, df_horizontal)

    def render_table(self, parent, df):
        if df is None or df.empty: return
        
        sf = ctk.CTkScrollableFrame(parent)
        sf.pack(fill="both", expand=True)
        
        # Headers
        headers = list(df.columns)
        for col, header in enumerate(headers):
            ctk.CTkLabel(sf, text=header, font=ctk.CTkFont(weight="bold"), text_color="#111827").grid(row=0, column=col, padx=8, pady=6)
            
        # Rows
        for r, row in df.iterrows():
            for c, col in enumerate(headers):
                header_name = headers[c]
                val = row[col]

                # Handle missing
                if val is None or pd.isna(val):
                    display = '-'
                else:
                    # Numeric formatting
                    if isinstance(val, numbers.Number):
                        # Percentage columns
                        if ('%' in header_name) or ('Vertical' in header_name):
                            display = f"{val:.2f}%"
                        # Currency / accounting format for monto-like fields
                        elif ('Saldo' in header_name) or ('Monto' in header_name):
                            try:
                                num = float(val)
                                if num < 0:
                                    display = f"(${abs(num):,.2f})"
                                else:
                                    display = f"${num:,.2f}"
                            except Exception:
                                display = str(val)
                        else:
                            display = str(val)
                    else:
                        display = str(val)


                ctk.CTkLabel(sf, text=display, text_color="#374151").grid(row=r+1, column=c, padx=8, pady=4)

    def export_report(self):
        """Export all financial data to Excel file"""
        try:
            from tkinter import filedialog, messagebox
            from datetime import datetime
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Ask user where to save
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"Reporte_Financiero_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not filename:
                return
            
            # Create Excel writer
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Fetch data
                income_accounts = CuentaDAO.obtener_cuentas_resultados_plataforma()
                balance_accounts = CuentaDAO.obtener_cuentas_balance_plataforma()
                
                # Helper functions
                import unicodedata
                def normalize(name):
                    normalized = unicodedata.normalize('NFD', name)
                    normalized = ''.join(c for c in normalized if unicodedata.category(c) != 'Mn')
                    return normalized.lower().strip()
                
                def get_account_value(accounts_list, account_name):
                    target_normalized = normalize(account_name)
                    for acc in accounts_list:
                        if normalize(acc.nombre_cuenta) == target_normalized:
                            return float(acc.saldo_actual)
                    return 0.0

                # --- 1. BALANCE GENERAL ---
                # Previous Period Data (Hardcoded matching load_balance_sheet)
                prev_balance_data = [
                    {"Cuenta": "ACTIVOS", "Monto": None},
                    {"Cuenta": "Efectivo", "Monto": 0},
                    {"Cuenta": "Mobiliario y accesorios", "Monto": 36000},
                    {"Cuenta": "Vehículos", "Monto": 50000},
                    {"Cuenta": "Depreciación de mobiliario y accesorios", "Monto": 0},
                    {"Cuenta": "Depreciación de vehículos", "Monto": 0},
                    {"Cuenta": "PASIVOS", "Monto": None},
                    {"Cuenta": "PATRIMONIO", "Monto": None},
                    {"Cuenta": "Capital social", "Monto": 86000},
                    {"Cuenta": "Utilidades retenidas", "Monto": 0}
                ]
                
                # Current Period Data & Logic matching load_balance_sheet
                curr_balance_map = {acc.nombre_cuenta: float(acc.saldo_actual) for acc in balance_accounts}
                
                # Calculate Totals exactly as in load_balance_sheet for Vertical Analysis
                curr_activos = 0
                for name, val in curr_balance_map.items():
                    if name in ["Efectivo", "Mobiliario y accesorios", "Vehiculos", "Vehículos"]:
                        curr_activos += val
                    elif "Depreciacion" in name or "Depreciación" in name:
                        curr_activos -= val # Assuming depreciation is positive in DB but subtracts
                
                prev_total_assets = 86000 # Hardcoded total from load_balance_sheet
                
                # Build Complete Data List
                balance_export = []
                
                # Combine Previous and Current keys (Normalized check to avoid duplicates)
                all_balance_keys = []
                processed_keys_norm = set()
                
                for item in prev_balance_data:
                    all_balance_keys.append(item["Cuenta"])
                    processed_keys_norm.add(normalize(item["Cuenta"]))
                    
                for k in curr_balance_map.keys():
                    if normalize(k) not in processed_keys_norm:
                        all_balance_keys.append(k)
                        processed_keys_norm.add(normalize(k))
                
                for key in all_balance_keys:
                    # Previous
                    prev_item = next((item for item in prev_balance_data if item["Cuenta"] == key), None)
                    val_prev = prev_item["Monto"] if prev_item else 0.0
                    if prev_item and prev_item["Monto"] is None: val_prev = None # Header
                    
                    # Current
                    val_curr = 0.0
                    if key in curr_balance_map:
                        val_curr = curr_balance_map[key]
                    else:
                        # Try normalized/fuzzy match
                        target_norm = normalize(key)
                        found = False
                        for db_key, db_val in curr_balance_map.items():
                            if normalize(db_key) == target_norm:
                                val_curr = db_val
                                found = True
                                break
                        # Specific fallback for Vehículos if normalization doesn't catch it (though it should)
                        if not found and key == "Vehículos" and "Vehiculos" in curr_balance_map:
                            val_curr = curr_balance_map["Vehiculos"]
                            
                    if prev_item and prev_item["Monto"] is None: val_curr = None # Header
                    
                    # Vertical Analysis
                    vert_prev = (val_prev / prev_total_assets * 100) if val_prev is not None and prev_total_assets else 0.0
                    vert_curr = (val_curr / curr_activos * 100) if val_curr is not None and curr_activos else 0.0
                    
                    if val_prev is None: # Header row
                        balance_export.append({
                            "Cuenta": key, "Año Base": None, "Vertical % Base": None, "Año Actual": None, "Vertical % Actual": None
                        })
                    else:
                        balance_export.append({
                            "Cuenta": key,
                            "Año Base": val_prev,
                            "Vertical % Base": vert_prev,
                            "Año Actual": val_curr,
                            "Vertical % Actual": vert_curr
                        })
                        
                df_balance = pd.DataFrame(balance_export)
                df_balance.to_excel(writer, sheet_name='Balance General', index=False)

                # --- 2. ESTADO DE RESULTADOS ---
                # Previous Period Data
                prev_income_data = [
                    {"Cuenta": "Ingresos", "Monto": 12800},
                    {"Cuenta": "Gasto de envío", "Monto": 2800},
                    {"Cuenta": "Renta", "Monto": 10000},
                    {"Cuenta": "Utilidad Antes de IR", "Monto": 3000},
                    {"Cuenta": "IR", "Monto": 0},
                    {"Cuenta": "Utilidad del período", "Monto": 0}
                ]
                
                # Current Period Data
                curr_income_map = {}
                curr_ingresos = 0
                curr_gastos = 0
                for acc in income_accounts:
                    val = float(acc.saldo_actual)
                    curr_income_map[acc.nombre_cuenta] = val
                    if acc.id_tipo_cuenta == 6: curr_ingresos += val
                    elif acc.id_tipo_cuenta == 7: curr_gastos += val
                
                # Calculated fields for current
                curr_util_antes = curr_ingresos - curr_gastos
                curr_income_map["Ingresos"] = curr_ingresos
                curr_income_map["Utilidad Antes de IR"] = curr_util_antes
                curr_income_map["IR"] = 0
                curr_income_map["Utilidad del período"] = curr_util_antes 
                
                # Build Export List
                income_export = []
                prev_ingresos_total = 12800
                
                all_income_keys = []
                processed_income_keys_norm = set()
                
                for item in prev_income_data:
                    all_income_keys.append(item["Cuenta"])
                    processed_income_keys_norm.add(normalize(item["Cuenta"]))
                    
                for k in curr_income_map.keys():
                    if normalize(k) not in processed_income_keys_norm:
                        all_income_keys.append(k)
                        processed_income_keys_norm.add(normalize(k))
                    
                for key in all_income_keys:
                    # Previous
                    prev_item = next((item for item in prev_income_data if item["Cuenta"] == key), None)
                    val_prev = prev_item["Monto"] if prev_item else 0.0
                    
                    # Current
                    val_curr = curr_income_map.get(key, 0.0)
                    if key not in curr_income_map:
                         # Try normalized match
                        target_norm = normalize(key)
                        for db_key, db_val in curr_income_map.items():
                            if normalize(db_key) == target_norm:
                                val_curr = db_val
                                break
                    
                    # Vertical
                    vert_prev = (val_prev / prev_ingresos_total * 100) if prev_ingresos_total else 0.0
                    vert_curr = (val_curr / curr_ingresos * 100) if curr_ingresos else 0.0
                    
                    income_export.append({
                        "Cuenta": key,
                        "Año Base": val_prev,
                        "Vertical % Base": vert_prev,
                        "Año Actual": val_curr,
                        "Vertical % Actual": vert_curr
                    })
                    
                df_income = pd.DataFrame(income_export)
                df_income.to_excel(writer, sheet_name='Estado de Resultados', index=False)

                # --- 3. ANÁLISIS HORIZONTAL (Separado) ---
                
                # 3.1 Balance Horizontal
                horizontal_balance_data = []
                for row in balance_export:
                    if row["Año Base"] is not None:
                        val_prev = row["Año Base"]
                        val_curr = row["Año Actual"]
                        var_abs = val_curr - val_prev
                        var_pct = (var_abs / val_prev * 100) if val_prev != 0 else 0.0
                        
                        horizontal_balance_data.append({
                            "Cuenta": row["Cuenta"],
                            "Año Base": val_prev,
                            "Año Actual": val_curr,
                            "Variación $": var_abs,
                            "Variación %": var_pct
                        })
                df_horiz_bal = pd.DataFrame(horizontal_balance_data)
                df_horiz_bal.to_excel(writer, sheet_name='Análisis Horiz. Balance', index=False)
                
                # 3.2 Income Horizontal
                horizontal_income_data = []
                for row in income_export:
                    val_prev = row["Año Base"]
                    val_curr = row["Año Actual"]
                    var_abs = val_curr - val_prev
                    var_pct = (var_abs / val_prev * 100) if val_prev != 0 else 0.0
                    
                    horizontal_income_data.append({
                        "Cuenta": row["Cuenta"],
                        "Año Base": val_prev,
                        "Año Actual": val_curr,
                        "Variación $": var_abs,
                        "Variación %": var_pct
                    })
                df_horiz_inc = pd.DataFrame(horizontal_income_data)
                df_horiz_inc.to_excel(writer, sheet_name='Análisis Horiz. Resultados', index=False)

                # --- 4. ORIGEN Y APLICACIÓN (Logic from load_cash_flow) ---
                sources_data = []
                uses_data = []
                
                # Re-using prev_balance_data structure but adding 'Tipo' for logic
                # Mapping from load_cash_flow
                prev_oa_data = [
                    {"Cuenta": "Efectivo", "Monto": 0, "Tipo": "Activo"},
                    {"Cuenta": "Mobiliario y accesorios", "Monto": 36000, "Tipo": "Activo"},
                    {"Cuenta": "Vehículos", "Monto": 50000, "Tipo": "Activo"},
                    {"Cuenta": "Depreciación de mobiliario y accesorios", "Monto": 0, "Tipo": "Activo"},
                    {"Cuenta": "Depreciación de vehículos", "Monto": 0, "Tipo": "Activo"},
                    {"Cuenta": "Capital Social", "Monto": 86000, "Tipo": "Capital"},
                    {"Cuenta": "Utilidades Retenidas", "Monto": 0, "Tipo": "Capital"}
                ]
                
                for row in prev_oa_data:
                    cuenta = row["Cuenta"]
                    monto_prev = row["Monto"]
                    tipo = row["Tipo"]
                    
                    # Find current value
                    monto_curr = 0.0
                    
                    # Try exact match first
                    if cuenta in curr_balance_map:
                        monto_curr = curr_balance_map[cuenta]
                    else:
                        # Try normalized matching
                        target_norm = normalize(cuenta)
                        for db_key, db_val in curr_balance_map.items():
                            if normalize(db_key) == target_norm:
                                monto_curr = db_val
                                break
                    
                    diff = monto_curr - monto_prev
                    
                    if diff == 0 or abs(diff) < 0.01:
                        continue
                        
                    # Logic:
                    # Activo: Decrease -> Source, Increase -> Use
                    # Pasivo/Capital: Increase -> Source, Decrease -> Use
                    
                    if tipo == "Activo":
                        if diff < 0:
                            sources_data.append({"Cuenta": cuenta, "Monto": abs(diff)})
                        elif diff > 0:
                            uses_data.append({"Cuenta": cuenta, "Monto": abs(diff)})
                    else: # Pasivo or Capital
                        if diff > 0:
                            sources_data.append({"Cuenta": cuenta, "Monto": abs(diff)})
                        elif diff < 0:
                            uses_data.append({"Cuenta": cuenta, "Monto": abs(diff)})

                df_sources = pd.DataFrame(sources_data)
                df_uses = pd.DataFrame(uses_data)
                
                # Write to single sheet with headers
                writer.sheets['Origen y Aplicación'] = writer.book.create_sheet('Origen y Aplicación')
                sheet = writer.sheets['Origen y Aplicación']
                
                sheet.cell(row=1, column=1, value="FUENTES (ORIGEN)")
                df_sources.to_excel(writer, sheet_name='Origen y Aplicación', index=False, startrow=1)
                
                start_row_uses = len(df_sources) + 4
                sheet.cell(row=start_row_uses, column=1, value="USOS (APLICACIÓN)")
                df_uses.to_excel(writer, sheet_name='Origen y Aplicación', index=False, startrow=start_row_uses)


                # --- 5. RAZONES FINANCIERAS ---
                if hasattr(self, 'ratios_data') and self.ratios_data:
                    df_ratios = pd.DataFrame(self.ratios_data)
                    df_ratios.to_excel(writer, sheet_name='Razones Financieras', index=False)

                # --- 6. FLUJOS DE EFECTIVO (Directo e Indirecto) ---
                # Re-calculate values exactly as in load_cash_flow_statements
                ingresos = get_account_value(income_accounts, "Ingresos")
                gasto_envio = get_account_value(income_accounts, "Gasto de envio")
                renta = get_account_value(income_accounts, "Renta")
                
                efectivo_actual = get_account_value(balance_accounts, "Efectivo")
                mobiliario = get_account_value(balance_accounts, "Mobiliario y accesorios")
                vehiculos = get_account_value(balance_accounts, "Vehiculos")
                dep_mobiliario = get_account_value(balance_accounts, "Depreciacion de mobiliario y accesorios")
                dep_vehiculos = get_account_value(balance_accounts, "Depreciacion de vehiculos")
                capital_social = get_account_value(balance_accounts, "Capital social")
                
                total_gastos = gasto_envio + renta
                utilidad_neta = ingresos - total_gastos
                
                efectivo_inicial = 0
                mobiliario_inicial = 36000
                vehiculos_inicial = 50000
                
                compra_mobiliario = max(0, mobiliario - mobiliario_inicial)
                compra_vehiculos = max(0, vehiculos - vehiculos_inicial)
                total_compra_activos = compra_mobiliario + compra_vehiculos
                aportacion_capital = max(0, capital_social - 86000)
                depreciacion_total = dep_mobiliario + dep_vehiculos
                
                # Direct Method
                cobros_clientes = ingresos
                pagos_operativos = -(gasto_envio + renta)
                efectivo_operacion_directo = cobros_clientes + pagos_operativos
                aumento_neto_efectivo = efectivo_operacion_directo - total_compra_activos + aportacion_capital

                direct_rows = [
                    {"Concepto": "MÉTODO DIRECTO", "Monto": ""},
                    {"Concepto": "ACTIVIDADES DE OPERACIÓN", "Monto": ""},
                    {"Concepto": "Cobros de clientes (Ingresos)", "Monto": cobros_clientes},
                    {"Concepto": "Pago de gasto de envío", "Monto": -gasto_envio},
                    {"Concepto": "Pago de renta", "Monto": -renta},
                    {"Concepto": "Efectivo neto de actividades de operación", "Monto": efectivo_operacion_directo},
                    {"Concepto": "", "Monto": ""},
                    {"Concepto": "ACTIVIDADES DE INVERSIÓN", "Monto": ""},
                    {"Concepto": "Compra de mobiliario y accesorios", "Monto": -compra_mobiliario if compra_mobiliario > 0 else 0},
                    {"Concepto": "Compra de vehículos", "Monto": -compra_vehiculos if compra_vehiculos > 0 else 0},
                    {"Concepto": "Efectivo neto de actividades de inversión", "Monto": -total_compra_activos},
                    {"Concepto": "", "Monto": ""},
                    {"Concepto": "ACTIVIDADES DE FINANCIAMIENTO", "Monto": ""},
                    {"Concepto": "Aportaciones de capital", "Monto": aportacion_capital},
                    {"Concepto": "Pago de dividendos", "Monto": 0},
                    {"Concepto": "Efectivo neto de actividades de financiamiento", "Monto": aportacion_capital},
                    {"Concepto": "", "Monto": ""},
                    {"Concepto": "Aumento (disminución) neto en efectivo", "Monto": aumento_neto_efectivo},
                    {"Concepto": "Efectivo al inicio del período", "Monto": efectivo_inicial},
                    {"Concepto": "Efectivo al final del período", "Monto": efectivo_inicial + aumento_neto_efectivo}
                ]
                
                # Indirect Method
                efectivo_operacion_indirecto = utilidad_neta + depreciacion_total
                aumento_neto_efectivo_indirecto = efectivo_operacion_indirecto - total_compra_activos + aportacion_capital
                
                indirect_rows = [
                    {"Concepto": "MÉTODO INDIRECTO", "Monto": ""},
                    {"Concepto": "ACTIVIDADES DE OPERACIÓN", "Monto": ""},
                    {"Concepto": "Utilidad neta", "Monto": utilidad_neta},
                    {"Concepto": "Ajustes para conciliar utilidad neta:", "Monto": ""},
                    {"Concepto": "  Depreciación de mobiliario", "Monto": dep_mobiliario},
                    {"Concepto": "  Depreciación de vehículos", "Monto": dep_vehiculos},
                    {"Concepto": "Efectivo neto de actividades de operación", "Monto": efectivo_operacion_indirecto},
                    {"Concepto": "", "Monto": ""},
                    {"Concepto": "ACTIVIDADES DE INVERSIÓN", "Monto": ""},
                    {"Concepto": "Compra de mobiliario y accesorios", "Monto": -compra_mobiliario if compra_mobiliario > 0 else 0},
                    {"Concepto": "Compra de vehículos", "Monto": -compra_vehiculos if compra_vehiculos > 0 else 0},
                    {"Concepto": "Efectivo neto de actividades de inversión", "Monto": -total_compra_activos},
                    {"Concepto": "", "Monto": ""},
                    {"Concepto": "ACTIVIDADES DE FINANCIAMIENTO", "Monto": ""},
                    {"Concepto": "Aportaciones de capital", "Monto": aportacion_capital},
                    {"Concepto": "Pago de dividendos", "Monto": 0},
                    {"Concepto": "Efectivo neto de actividades de financiamiento", "Monto": aportacion_capital},
                    {"Concepto": "", "Monto": ""},
                    {"Concepto": "Aumento (disminución) neto en efectivo", "Monto": aumento_neto_efectivo_indirecto},
                    {"Concepto": "Efectivo al inicio del período", "Monto": efectivo_inicial},
                    {"Concepto": "Efectivo al final del período", "Monto": efectivo_inicial + aumento_neto_efectivo_indirecto}
                ]
                
                df_direct = pd.DataFrame(direct_rows)
                df_indirect = pd.DataFrame(indirect_rows)
                
                # Write side-by-side or stacked. Stacked is easier for Excel readability if columns differ, but here they are same.
                # Let's put them in one sheet, separated by columns.
                df_direct.to_excel(writer, sheet_name='Flujos de Efectivo', index=False, startcol=0)
                df_indirect.to_excel(writer, sheet_name='Flujos de Efectivo', index=False, startcol=3)

                # --- 7. PROFORMA (Si existe) ---
                if hasattr(self, 'proforma_bs_df') and hasattr(self, 'proforma_is_df'):
                    # Write Balance Sheet Proforma
                    self.proforma_bs_df.to_excel(writer, sheet_name='Proforma', index=False, startcol=0)
                    writer.sheets['Proforma'].cell(row=1, column=1, value="Balance General Proforma")
                    
                    # Write Income Statement Proforma (side-by-side)
                    # Calculate start column (len of BS columns + gap)
                    start_col_is = len(self.proforma_bs_df.columns) + 2
                    self.proforma_is_df.to_excel(writer, sheet_name='Proforma', index=False, startcol=start_col_is)
                    writer.sheets['Proforma'].cell(row=1, column=start_col_is+1, value="Estado de Resultados Proforma")

            messagebox.showinfo("Éxito", f"Reporte exportado exitosamente a:\n{filename}")
            
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Error", f"Error al exportar reporte:\n{str(e)}")
            import traceback
            traceback.print_exc()
